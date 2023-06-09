import os

from openpyxl import Workbook, load_workbook


def main():
    allList = []
    keyword = "TestJob"
    current_dir = os.path.dirname(os.path.abspath(__file__))
    fin_path = os.path.join(current_dir, "ALLExcel")

    for file in os.listdir(fin_path):
        if file.endswith(".xlsx"):
            file_path = os.path.join(fin_path, file)
            wb = load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.values:
                    for value in row:
                        if keyword == value:
                            print(sheet)
                            print(row)
                            allList.append(row)


    print(allList)
    wbook = Workbook()
    wba = wbook.active

    for array in allList:
        wba.append(array)

    wbook.save("MySheet.xlsx")




main()